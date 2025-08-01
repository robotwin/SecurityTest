#定义合并单元格的函数    
def Merge_cells(ws,target_list,start_row,col):
    '''
    ws: 是需要操作的工作表
    target_list: 是目标列表，即含有重复数据的列表
    start_row: 是开始行，即工作表中开始比对数据的行（需要将标题除开）
    col: 是需要处理数据的列
    '''
    print("hey siri")
    start = 0 #开始行计数，初试值为0，对应列表中的第1个元素的位置0
    end = 0 #结束行计数，初试值为0，对应列表中的第1个元素的位置0
    reference = target_list[0] #设定基准，以列表中的第一个字符串开始
    for i in range(len(target_list)): #遍历列表
        if target_list[i] != reference: #开始比对，如果内容不同执行如下
            reference = target_list[i] #基准变成列表中下一个字符串
            end = i - 1 #列计数器
            ws.merge_cells(col + str(start + start_row) + ":"+col + str(end + start_row))
            start = end + 1
        if i == len(target_list) - 1: #遍历到最后一行，按如下操作
            end = i
            ws.merge_cells(col + str(start + start_row) + ":"+ col + str(end + start_row))

# from openpyxl import load_workbook意思是从'openpyxl'库中导入'load_workbook'函数,
# 这个函数是'openpyxl'中用来加载现有Excel文件主要函数,使用这个可以打开一个已经存储的Excel文件,
# 以便后续对其进行读取,修改或者其他操作.
#获取Excel表格中的数据
from openpyxl import load_workbook #用于读取Excel中的信息
# 加载Excel文件,获取所有工作表的名称列表,一旦加载了工作簿,你可以访问工作簿中的工作表,单元格内容等.
wb = load_workbook('产品清单.xlsx')
# 返回一个包含所有工作表名称的列表.
sheet_names = wb.get_sheet_names()
# 是一个for循环语句,用于遍历一个列表sheet_names中的每个元素,并将每个元素依次赋值给变量sheet_name
for sheet_name in sheet_names: #遍历每个工作表，抓取数据，并根据要求合并单元格
    ws = wb[sheet_name]
    customer_list = [] #客户名称
    pn_list = [] #产品编码

    for row in range(6,ws.max_row-2):
        customer = ws['B' + str(row)].value
        pn = ws['C' + str(row)].value
        customer_list.append(customer)
        pn_list.append(pn)
        
    #调用以上定义的合并单元格函数`Merge_cells`做单元格合并操作    
    start_row=6 #开始行是第六行
    Merge_cells(ws,customer_list,start_row,"B") #"B" - 客户名称是在B列
    Merge_cells(ws,pn_list,start_row,"C") #"C" - 产品编码是在C列
        
wb.save("产品清单-合并单元.xlsx")